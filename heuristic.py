#!/usr/bin/env python3
"""
Heuristic Algorithm for Double Round-Robin Scheduling (Rest Mismatch Problem)
for 18 teams (Euroleague 2025 example).

For n = 18:
  - A single round-robin has n-1 = 17 rounds, each with 9 games.
  - Each round is split into two periods:
       Day 1: first ceil(9/2) = 5 games
       Day 2: remaining 4 games.
       
Heuristic Step 1:
  Partition the 17 rounds into three stages:
    • Stage 1: rounds 1 to 8
    • Stage 2: rounds 9 to 12
    • Stage 3: rounds 13 to 17

Heuristic Step 2:
  Part 1:
    - Swap round 1 with round 8.
    - In the new round 8, swap the first block (block_size = 2) of Day 1 with the last block of Day 2.
  Part 2:
    - Swap round 9 with round 13.
    - In the new round 13, swap the first block of odd-indexed games of Day 1 with the first block of even-indexed games of Day 2.

Mirroring:
  The final single round-robin schedule (17 rounds) is mirrored by reversing each match to create rounds 18 to 34.
"""

import math
import pandas as pd
import copy
import random 
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime, timedelta

def save_schedule_to_excel(schedule, filename="schedule.xlsx"):
    """
    Converts the schedule dictionary into a DataFrame format and saves it to an Excel file.
    
    - Periods (1, 2) become columns.
    - Rounds become rows.
    """
    formatted_schedule = {}

    for r in sorted(schedule.keys()):
        formatted_schedule[r] = {
            "Period 1": ", ".join(f"{m[0]}-{m[1]}" for m in schedule[r]["Day 1"]),
            "Period 2": ", ".join(f"{m[0]}-{m[1]}" for m in schedule[r]["Day 2"])
        }

    # Convert to DataFrame
    schedule_df = pd.DataFrame.from_dict(formatted_schedule, orient="index")
    schedule_df.index.name = "Round"

    # Save to Excel
    schedule_df.to_excel(filename)

    return schedule_df
  
  
import math

def generate_initial_schedule(n):
    """
    Generate a single round-robin schedule using the circle method.
    For n teams, there are n-1 rounds with n/2 games per round.
    Then split each round into two periods:
       Day 1: first ceil(n/2 / 2) games (for n=18, ceil(9/2)=5)
       Day 2: the remaining games.
    Returns:
      schedule[r] = {"Day 1": [match, ...], "Day 2": [match, ...]}
    where each match is a tuple (team1, team2).
    """
    num_rounds = n - 1
    teams = list(range(1, n+1))
    rounds_list = []
    
    # Standard circle method: fix the first team, rotate the others.
    for r in range(num_rounds):
        round_matches = []
        for i in range(n // 2):
            t1 = teams[i]
            t2 = teams[n - 1 - i]
            round_matches.append((t1, t2))
        rounds_list.append(round_matches)
        # Rotate: fix teams[0], rotate remaining clockwise.
        teams = [teams[0]] + [teams[-1]] + teams[1:-1]
    
    # Split each round into two periods.
    schedule = {}
    for r in range(num_rounds):
        matches = rounds_list[r]  # 9 matches per round for n=18.
        split_index = math.ceil(len(matches) / 2)  # For 9, split_index = 5.
        day1 = matches[:split_index]
        day2 = matches[split_index:]
        schedule[r+1] = {"Day 1": day1, "Day 2": day2}
    return schedule


def print_schedule(schedule, title="Schedule"):
    print(f"\n{title}:")
    for r in sorted(schedule.keys()):
        d1 = ", ".join(f"{m[0]}-{m[1]}" for m in schedule[r]["Day 1"])
        d2 = ", ".join(f"{m[0]}-{m[1]}" for m in schedule[r]["Day 2"])
        print(f"Round {r:2d}: Day 1: {d1:<50} | Day 2: {d2}")

def heuristic_step1(schedule):
    """
    Partition the 17 rounds (for n=18) into three stages.
    For example, we use:
      Stage 1: rounds 1 to 8,
      Stage 2: rounds 9 to 12,
      Stage 3: rounds 13 to 17.
    Returns the combined schedule (here we simply partition, without re-computation).
    """
    n_rounds = len(schedule)  # should be 17 for n=18.
    stage1 = {r: schedule[r] for r in range(1, 9)}
    stage2 = {r: schedule[r] for r in range(9, 13)}
    stage3 = {r: schedule[r] for r in range(13, n_rounds+1)}
    combined = {}
    combined.update(stage1)
    combined.update(stage2)
    combined.update(stage3)
    return combined


def track_swaps(schedule):
    """
    Tracks swapped matches in heuristic_step2.
    Returns a DataFrame containing the details of swapped rounds and games.
    """
    swap_data = []

    # --- Part 1: Swap round 1 and round 8 ---
    swap_data.append({"Swap Type": "Round Swap", "Round A": 1, "Round B": 8, "Swapped Matches": "Full round swap"})
    day1_r8_old = list(schedule[8]["Day 1"])  # Period 1 for round 8 has 5 matches.
    day2_r8_old = list(schedule[8]["Day 2"])  # Period 2 for round 8 has 4 matches.
    swap_day1 = day1_r8_old[:2]  # First block (2 matches) of Period 1
    swap_day2 = day2_r8_old[-2:]  # Last block (2 matches) of Period 2

    swap_data.append({"Swap Type": "Game Swap", "Round A": 8, "Round B": 8, 
                      "Swapped Matches": f"{swap_day1} ↔ {swap_day2}"})

    # --- Part 2: Swap round 9 and round 13 ---
    swap_data.append({"Swap Type": "Round Swap", "Round A": 9, "Round B": 13, "Swapped Matches": "Full round swap"})
    day1_r13_old = list(schedule[13]["Day 1"])  # Period 1 has 5 games.
    day2_r13_old = list(schedule[13]["Day 2"])  # Period 2 has 4 games.
    
    odd_day1 = [day1_r13_old[i] for i in range(len(day1_r13_old)) if i % 2 == 0]
    even_day2 = [day2_r13_old[i] for i in range(len(day2_r13_old)) if i % 2 == 1]

    swap_odd = odd_day1[:2]
    swap_even = even_day2[:2]

    swap_data.append({"Swap Type": "Game Swap", "Round A": 13, "Round B": 13, 
                      "Swapped Matches": f"{swap_odd} ↔ {swap_even}"})

    return pd.DataFrame(swap_data)


def track_break_swaps(original_schedule, optimized_schedule, tracked_rounds=[5, 10, 15, 20, 25, 30]):
    """
    Tracks swapped home/away assignments (break mismatch swaps).
    
    Args:
    original_schedule (dict): The schedule before break mismatch swaps.
    optimized_schedule (dict): The schedule after break mismatch swaps.
    tracked_rounds (list): Specific rounds to track swaps.

    Returns:
    pd.DataFrame: A DataFrame containing the details of break mismatch swaps.
    """
    swap_data = []

    for round_num in tracked_rounds:
        if round_num not in original_schedule or round_num not in optimized_schedule:
            continue  # Skip if the round doesn't exist

        before_day1 = original_schedule[round_num]["Day 1"]
        before_day2 = original_schedule[round_num]["Day 2"]

        after_day1 = optimized_schedule[round_num]["Day 1"]
        after_day2 = optimized_schedule[round_num]["Day 2"]

        swapped_day1 = []
        swapped_day2 = []

        # Compare matches before and after swap to track home/away flips
        for i in range(len(before_day1)):
            if before_day1[i] == (after_day1[i][1], after_day1[i][0]):  # Swapped order
                swapped_day1.append(f"{before_day1[i]} ↔ {after_day1[i]}")

        for i in range(len(before_day2)):
            if before_day2[i] == (after_day2[i][1], after_day2[i][0]):  # Swapped order
                swapped_day2.append(f"{before_day2[i]} ↔ {after_day2[i]}")

        if swapped_day1 or swapped_day2:
            swap_data.append({
                "Round": round_num,
                "Day 1 Swaps": ", ".join(swapped_day1) if swapped_day1 else "No swaps",
                "Day 2 Swaps": ", ".join(swapped_day2) if swapped_day2 else "No swaps"
            })

    return pd.DataFrame(swap_data)

def mirror_schedule(schedule):
    """
    Mirror each round immediately after itself.
    For each round r, create round r+1 by reversing each match (i, j) -> (j, i).
    This prevents long home/away streaks by alternating rounds.
    
    Returns a dictionary with mirrored rounds directly following the original rounds.
    """
    mirrored_schedule = {}
    new_round_index = 1

    for r in sorted(schedule.keys()):
        # Original round
        mirrored_schedule[new_round_index] = schedule[r]
        new_round_index += 1

        # Mirrored round (immediately after)
        mirror_round = {
            "Day 1": [(m[1], m[0]) for m in schedule[r]["Day 1"]],
            "Day 2": [(m[1], m[0]) for m in schedule[r]["Day 2"]]
        }
        mirrored_schedule[new_round_index] = mirror_round
        new_round_index += 1

    return mirrored_schedule


def extract_rounds(schedule, round_list):
    """
    Given a full schedule (dictionary of round -> {"Day 1": [...], "Day 2": [...]})
    and a list of round numbers, return a dict containing only those rounds.
    Each entry is round -> (period1_list, period2_list).
    """
    partial = {}
    for r in round_list:
        period1 = schedule[r]["Day 1"]
        period2 = schedule[r]["Day 2"]
        partial[r] = (period1, period2)
    return partial

def print_partial_table(round_data, step_label):
    """
    Prints a partial schedule in a table style, with two columns: Period 1 and Period 2.
    round_data is a dict: round -> (list_of_matches_day1, list_of_matches_day2).
    step_label is a string describing the step, e.g. '(a) Before swap of round 1 and round 8'.
    """
    print(step_label)
    print("Round   Day 1                        Day 2")
    for r in sorted(round_data.keys()):
        p1, p2 = round_data[r]
        p1_str = " ".join(f"{m[0]}-{m[1]}" for m in p1)
        p2_str = " ".join(f"{m[0]}-{m[1]}" for m in p2)
        print(f"{r:<6}  {p1_str:<30}  {p2_str}")
    print()


def heuristic_step2(schedule):
    """
    Apply swapping operations (Step 2 of the heuristic) for n=18, and
    print partial schedules (round 1 & 8, then 9 & 13) before and after each swap.
    Also tracks the swaps for logging.
    """
    new_schedule = copy.deepcopy(schedule)
    block_size = 2  # For n=18, we use block_size = 2.
    swap_records = []

    # --- Part 1: Swap round 1 with round 8 ---
    before_1_8 = extract_rounds(new_schedule, [1, 8])
    print_partial_table(before_1_8, "(a) Before swapping Round 1 and Round 8")

    # Swap the entire rounds
    new_schedule[1], new_schedule[8] = copy.deepcopy(new_schedule[8]), copy.deepcopy(new_schedule[1])

    # Then in the new Round 8, swap the first block_size games of Day 1 with the last block_size games of Day 2
    day1_r8 = list(new_schedule[8]["Day 1"])
    day2_r8 = list(new_schedule[8]["Day 2"])

    swap_day1 = day1_r8[:block_size]
    swap_day2 = day2_r8[-block_size:]
    new_schedule[8]["Day 1"] = swap_day2 + day1_r8[block_size:]
    new_schedule[8]["Day 2"] = day2_r8[:-block_size] + swap_day1

    # Track swap
    swap_records.append({"Swap Type": "Round Swap", "Round A": 1, "Round B": 8, "Swapped Matches": "Full round swap"})
    swap_records.append({"Swap Type": "Game Swap", "Round A": 8, "Round B": 8, "Swapped Matches": f"{swap_day1} ↔ {swap_day2}"})

    # 1B. Print partial table of Rounds 1 and 8 AFTER swap
    after_1_8 = extract_rounds(new_schedule, [1, 8])
    print_partial_table(after_1_8, "(b) After swapping Round 1 and Round 8")

    # --- Part 2: Swap round 9 with round 13 ---
    before_9_13 = extract_rounds(new_schedule, [9, 13])
    print_partial_table(before_9_13, "(c) Before swapping Round 9 and Round 13")

    # Swap the entire rounds
    new_schedule[9], new_schedule[13] = copy.deepcopy(new_schedule[13]), copy.deepcopy(new_schedule[9])

    # Then in the new Round 13, swap the first block_size odd-indexed games of Day 1 
    # with the first block_size even-indexed games of Day 2
    day1_r13 = list(new_schedule[13]["Day 1"])
    day2_r13 = list(new_schedule[13]["Day 2"])
    odd_day1 = [day1_r13[i] for i in range(len(day1_r13)) if i % 2 == 0]
    even_day2 = [day2_r13[i] for i in range(len(day2_r13)) if i % 2 == 1]
    swap_odd = odd_day1[:block_size]
    swap_even = even_day2[:block_size]

    new_day1_r13 = []
    new_day2_r13 = []
    odd_counter = 0
    even_counter = 0
    for i, game in enumerate(day1_r13):
        if i % 2 == 0 and odd_counter < block_size:
            new_day1_r13.append(swap_even[even_counter])
            even_counter += 1
            odd_counter += 1
        else:
            new_day1_r13.append(game)
    odd_counter = 0
    even_counter = 0
    for i, game in enumerate(day2_r13):
        if i % 2 == 1 and even_counter < block_size:
            new_day2_r13.append(swap_odd[odd_counter])
            odd_counter += 1
            even_counter += 1
        else:
            new_day2_r13.append(game)

    new_schedule[13]["Day 1"] = new_day1_r13
    new_schedule[13]["Day 2"] = new_day2_r13

    # Track swap
    swap_records.append({"Swap Type": "Round Swap", "Round A": 9, "Round B": 13, "Swapped Matches": "Full round swap"})
    swap_records.append({"Swap Type": "Game Swap", "Round A": 13, "Round B": 13, "Swapped Matches": f"{swap_odd} ↔ {swap_even}"})

    # 2B. Print partial table of Rounds 9 and 13 AFTER swap
    after_9_13 = extract_rounds(new_schedule, [9, 13])
    print_partial_table(after_9_13, "(d) After swapping Round 9 and Round 13")

    return new_schedule, pd.DataFrame(swap_records)


def minimize_breaks_with_tracking(schedule):
    """
    Adjust home/away assignments to reduce consecutive home/away games,
    tracking only a few key swaps to highlight.

    Args:
    schedule (dict): Dictionary of rounds -> { "Day 1": [matches], "Day 2": [matches] }

    Returns:
    dict: Optimized schedule with minimized breaks.
    """
    optimized_schedule = copy.deepcopy(schedule)  # Preserve original schedule
    swap_records = []  # Track **only key** swaps

    home_away_tracker = {team: {} for team in range(1, 19)}  # Track home/away per team per round

    # Step 1: Assign initial home/away status
    for round_num, round_data in optimized_schedule.items():
        for period, matches in round_data.items():
            for match in matches:
                t1, t2 = match
                home_away_tracker[t1][round_num] = "H"
                home_away_tracker[t2][round_num] = "A"

    # Step 2: Identify consecutive breaks
    consecutive_breaks = {}
    for team in range(1, 19):  # Teams are 1-18
        breaks = []
        last_status = None
        for round_num in range(1, 35):  # 34 rounds in DRR
            if round_num in home_away_tracker[team]:
                current_status = home_away_tracker[team][round_num]
                if last_status == current_status:  # Consecutive H or A
                    breaks.append(round_num)
                last_status = current_status
        if breaks:
            consecutive_breaks[team] = breaks

    # Step 3: Swap home/away assignments for **selected key swaps**
    key_swaps = []  # Store which rounds were swapped

    for team, break_rounds in consecutive_breaks.items():
        for break_round in break_rounds:
            if break_round < 34:  # Avoid out-of-bounds
                swap_round = break_round + 1

                # **Highlight only a few swaps**
                if break_round in [5, 10, 15, 20, 25, 30]:  # Track only these key swaps
                    key_swaps.append((break_round, swap_round))

                for period, matches in optimized_schedule[break_round].items():
                    for i, match in enumerate(matches):
                        if team in match:
                            t1, t2 = match
                            optimized_schedule[break_round][period][i] = (t2, t1) if t1 == team else (t1, t2)
                            break  # Ensure only one swap per match


    return optimized_schedule


def format_table(round_data, title):
    """
    Formats extracted round data into a DataFrame for display.

    Args:
    round_data (dict): Extracted round data (keys: round numbers, values: tuples of lists).

    Returns:
    pd.DataFrame: Formatted table.
    """
    formatted_data = []
    for round_num, days in round_data.items():
        day1_matches = " ".join(f"{m[0]}-{m[1]}" for m in days[0])  # Use tuple index 0 for Day 1
        day2_matches = " ".join(f"{m[0]}-{m[1]}" for m in days[1])  # Use tuple index 1 for Day 2

        formatted_data.append([round_num, day1_matches, day2_matches])

    return pd.DataFrame(formatted_data, columns=["Round", "Day 1", "Day 2"]).rename_axis(title)

  
def validate_schedule(schedule_excel):
  
    # Check if each team plays only once per round
    rounds_valid = True
    for round_number, row in schedule_excel.iterrows():
        teams_played = set()
        for period, matches in row.items():
            if pd.notna(matches):
                match_list = matches.split(", ")
                for match in match_list:
                    team1, team2 = map(int, match.split("-"))
                    if team1 in teams_played or team2 in teams_played:
                        rounds_valid = False
                        print(f"Error in Round {round_number}: Team {team1} or {team2} plays more than once.")
                    teams_played.add(team1)
                    teams_played.add(team2)
    
    # Check if all teams are included in every round
    all_teams = set(range(1, 19))
    rounds_complete = True
    for round_number, row in schedule_excel.iterrows():
        teams_in_round = set()
        for period, matches in row.items():
            if pd.notna(matches):
                match_list = matches.split(", ")
                for match in match_list:
                    team1, team2 = map(int, match.split("-"))
                    teams_in_round.add(team1)
                    teams_in_round.add(team2)
        if teams_in_round != all_teams:
            rounds_complete = False
            print(f"Error in Round {round_number}: Not all teams are present. Missing: {all_teams - teams_in_round}")
    
    # Print overall results
    if rounds_valid and rounds_complete:
        print("\n\nThe schedule is valid: Each team plays once per round, and all teams are present in each round.")
    else:
        print("The schedule has errors. Check the printed messages above.")


def fix_home_away_imbalance(schedule, n):
    """
    Identifies and swaps matches to fix home/away imbalance for teams with 18 home or 18 away games.
    The approach swaps a match where one team has an imbalance with another match where both teams are balanced.
    """
    home_away_count = {team: {"Home": 0, "Away": 0} for team in range(1, n+1)}

    # Count home and away matches per team
    for r in schedule:
        for period in ["Day 1", "Day 2"]:
            for match in schedule[r][period]:
                home_away_count[match[0]]["Home"] += 1
                home_away_count[match[1]]["Away"] += 1

    # Identify teams with imbalance
    imbalanced_teams = {team: data for team, data in home_away_count.items() if data["Home"] != 17}

    if len(imbalanced_teams) != 2:
        return schedule, imbalanced_teams  # No fix needed or more than 2 imbalanced teams

    team_a, team_b = list(imbalanced_teams.keys())

    # Find a round where the imbalance occurs and swap a match
    for r in schedule:
        for period in ["Day 1", "Day 2"]:
            matches = schedule[r][period]
            for i, (team1, team2) in enumerate(matches):
                if (team1 == team_a and home_away_count[team1]["Home"] > 17) or \
                   (team2 == team_a and home_away_count[team2]["Away"] > 17):
                    # Swap the match to correct imbalance
                    matches[i] = (team2, team1)
                    home_away_count[team1]["Home"] -= 1
                    home_away_count[team1]["Away"] += 1
                    home_away_count[team2]["Away"] -= 1
                    home_away_count[team2]["Home"] += 1
                    return schedule, imbalanced_teams  # Return after fixing

    return schedule, imbalanced_teams  # If no swap was found
def check_home_away_balance(schedule, n):
    """
    Check if any team has an unbalanced home/away distribution (i.e., 17 home or 17 away games).
    Returns a dictionary with teams as keys and their home/away count.
    """
    home_away_count = {team: {"Home": 0, "Away": 0} for team in range(1, n+1)}

    for r in schedule:
        for period in ["Day 1", "Day 2"]:
            for match in schedule[r][period]:
                home_away_count[match[0]]["Home"] += 1
                home_away_count[match[1]]["Away"] += 1

    # Convert results into a DataFrame for better visualization
    home_away_df = pd.DataFrame.from_dict(home_away_count, orient="index")
    home_away_df.index.name = "Team"

    return home_away_df

if __name__ == '__main__':
    n = 18
    print("Generating initial single round-robin schedule for 18 teams...")
    init_schedule = generate_initial_schedule(n)
    
    # Step 1: Apply Heuristic Step 1
    schedule_stage = heuristic_step1(init_schedule)

    # Step 2: Apply Heuristic Step 2 with tracking
    final_single, swapped_matches_df = heuristic_step2(schedule_stage)

    # Save swapped matches to Excel
    swapped_matches_filepath = "swapped_matches.xlsx"
    swapped_matches_df.to_excel(swapped_matches_filepath, index=False)

    # Step 3: Mirror the schedule for Double Round-Robin
    double_schedule = mirror_schedule(init_schedule)

    break_swap_table = track_break_swaps(init_schedule, double_schedule)
    break_swap_table.to_excel("break_swap_table.xlsx", index=False)
    
    # Step 5: Save the final schedule to Excel
    print("Break minimization applied. Optimized schedule saved as 'optimized_schedule.xlsx.")
    schedule_excel = save_schedule_to_excel(init_schedule, filename="optimized_schedule.xlsx")

    # Step 6: Validate the final schedule
    validate_schedule(schedule_excel)
    
    # Apply the fix for the home/away imbalance
    fixed_schedule, imbalanced_teams_after_fix = fix_home_away_imbalance(double_schedule, 18)
    
    # Check the home/away balance again
    home_away_balance_fixed_df = check_home_away_balance(fixed_schedule, 18)
    #===================================================================
    
    
    # Re-define team mappings
    num_to_team = {
        i + 1: team for i, team in enumerate([
            "ALBA BERLIN", "ANADOLU EFES ISTANBUL", "AS MONACO", "BASKONIA VITORIA-GASTEIZ",
            "CRVENA ZVEZDA MERIDIANBET BELGRADE", "EA7 EMPORIO ARMANI MILAN", "FC BARCELONA", "FC BAYERN MUNICH",
            "FENERBAHCE BEKO ISTANBUL", "LDLC ASVEL VILLEURBANNE", "MACCABI PLAYTIKA TEL AVIV",
            "OLYMPIACOS PIRAEUS", "PANATHINAIKOS AKTOR ATHENS", "PARIS BASKETBALL", "PARTIZAN MOZZART BET BELGRADE",
            "REAL MADRID", "VIRTUS SEGAFREDO BOLOGNA", "ZALGIRIS KAUNAS"
        ])
    }
    
    abbreviation_map = {
        "ALBA BERLIN": "ALB",
        "ANADOLU EFES ISTANBUL": "AEF",
        "AS MONACO": "MON",
        "BASKONIA VITORIA-GASTEIZ": "BAS",
        "CRVENA ZVEZDA MERIDIANBET BELGRADE": "CZV",
        "EA7 EMPORIO ARMANI MILAN": "MIL",
        "FC BARCELONA": "BAR",
        "FC BAYERN MUNICH": "BAY",
        "FENERBAHCE BEKO ISTANBUL": "FEN",
        "LDLC ASVEL VILLEURBANNE": "ASV",
        "MACCABI PLAYTIKA TEL AVIV": "MAC",
        "OLYMPIACOS PIRAEUS": "OLY",
        "PANATHINAIKOS AKTOR ATHENS": "PAO",
        "PARIS BASKETBALL": "PSG",
        "PARTIZAN MOZZART BET BELGRADE": "PAR",
        "REAL MADRID": "RMA",
        "VIRTUS SEGAFREDO BOLOGNA": "VIR",
        "ZALGIRIS KAUNAS": "ZAL"
    }
    
    
    
    # Convert schedule to DataFrame with abbreviations and full names
    def convert_schedule(schedule, num_to_team, abbreviation_map):
        schedule_abbr = {}
        schedule_full = {}
    
        for round_num, matches in schedule.items():
            schedule_abbr[round_num] = {
                "Saturday": ", ".join(f"{abbreviation_map[num_to_team[m[0]]]} - {abbreviation_map[num_to_team[m[1]]]}" for m in matches["Day 1"]),
                "Sunday": ", ".join(f"{abbreviation_map[num_to_team[m[0]]]} - {abbreviation_map[num_to_team[m[1]]]}" for m in matches["Day 2"])
            }
    
            schedule_full[round_num] = {
                "Saturday": ", ".join(f"{num_to_team[m[0]]} - {num_to_team[m[1]]}" for m in matches["Day 1"]),
                "Sunday": ", ".join(f"{num_to_team[m[0]]} - {num_to_team[m[1]]}" for m in matches["Day 2"])
            }
    
        return pd.DataFrame.from_dict(schedule_abbr, orient="index"), pd.DataFrame.from_dict(schedule_full, orient="index")
    
    # Convert the finalized schedule
    schedule_abbr_df, schedule_full_df = convert_schedule(fixed_schedule, num_to_team, abbreviation_map)
    
    # Convert abbreviation map to DataFrame
    abbrev_df = pd.DataFrame(list(abbreviation_map.items()), columns=["Team Name", "Abbreviation"])
    
    
    # Extracted schedule dates from the official calendar
    official_schedule_dates = {
        1: "3-4 October, 2024", 2: "10-11 October, 2024", 3: "15-16 October, 2024",
        4: "17-18 October, 2024", 5: "24-25 October, 2024", 6: "29-30 October, 2024",
        7: "31 October -1 November, 2024", 8: "7-8 November, 2024", 9: "12-13 November, 2024",
        10: "14-15 November, 2024", 11: "21-22 November, 2024", 12: "28-29 November, 2024",
        13: "3-4 December, 2024", 14: "5-6 December, 2024", 15: "12-13 December, 2024",
        16: "17-18 December, 2024", 17: "19-20 December, 2024", 18: "26-27 December, 2024",
        19: "2-3 January, 2025", 20: "9-10 January, 2025", 21: "14-15 January, 2025",
        22: "16-17 January, 2025", 23: "23-24 January, 2025", 24: "30-31 January, 2025",
        25: "4-5 February, 2025", 26: "6-7 February, 2025", 27: "27-28 February, 2025",
        28: "6-7 March, 2025", 29: "13-14 March, 2025", 30: "20-21 March, 2025",
        31: "25-26 March, 2025", 32: "27-28 March, 2025", 33: "3-4 April, 2025",
        34: "10-11 April, 2025"
    }
    
    
    
    # Convert the original schedule dates to fall on Saturday and Sunday
    def adjust_to_weekend(date_str):
        """Takes a date range (e.g., '3-4 October, 2024') and returns the closest Saturday/Sunday dates."""
        parts = date_str.split(",")[0].split(" ")  # Extract day(s) and month
        days = parts[0].split("-")  # Handle single or ranged days
        month = parts[1]  # Extract the month
        year = int(date_str[-4:])  # Extract year
    
        # Convert the first day in the range to a datetime object
        first_day = int(days[0])
        date_obj = datetime.strptime(f"{first_day} {month} {year}", "%d %B %Y")
    
        # Find the next Saturday/Sunday
        while date_obj.weekday() not in [5, 6]:  # 5 = Saturday, 6 = Sunday
            date_obj += timedelta(days=1)
    
        # Get the next day's match on the other weekend day
        second_date = date_obj + timedelta(days=1) if date_obj.weekday() == 5 else date_obj - timedelta(days=1)
    
        return f"{date_obj.strftime('%d %B, %Y')} - {second_date.strftime('%d %B, %Y')}"
    
    # Adjust all rounds to weekend dates
    adjusted_schedule_dates = {round_num: adjust_to_weekend(date) for round_num, date in official_schedule_dates.items()}
    
    
    # Adjust the schedule DataFrames to use dates instead of round numbers
    
    # Convert the round numbers in the schedule DataFrames to official dates (adjusted to Saturday/Sunday)
    schedule_abbr_df.index = schedule_abbr_df.index.map(lambda r: adjusted_schedule_dates.get(r, f"Round {r}"))
    schedule_full_df.index = schedule_full_df.index.map(lambda r: adjusted_schedule_dates.get(r, f"Round {r}"))
    
    # Reset the index and place the "Date" column right after the index
    schedule_abbr_df.reset_index(inplace=True)
    schedule_full_df.reset_index(inplace=True)
    
    # Rename the index column to "Round"
    schedule_abbr_df.rename(columns={"index": "Date"}, inplace=True)
    schedule_full_df.rename(columns={"index": "Date"}, inplace=True)
    
    
    # Save the updated schedule DataFrames to an Excel file
    with pd.ExcelWriter("Euroleague_Schedule_2025.xlsx") as writer:
        # Schedule with abbreviations
        schedule_abbr_df.to_excel(writer, sheet_name="Schedule_Abbreviations")
        
        # Schedule with full team names
        schedule_full_df.to_excel(writer, sheet_name="Schedule_FullNames")
        
        # Abbreviations mapping
        abbrev_df.to_excel(writer, sheet_name="Abbreviations", index=False)
    

    
    def validate_schedule(file_path):
        wb = load_workbook(file_path)
        abbrev_sheet = wb['Abbreviations']
        
        # Load all 18 teams
        teams = {}
        for row in abbrev_sheet.iter_rows(min_row=2):
            teams[row[1].value] = row[0].value  # {abbreviation: full_name}
    
        home = defaultdict(int)
        away = defaultdict(int)
    
        # Process games
        schedule_sheet = wb['Schedule_Abbreviations']
        for row in schedule_sheet.iter_rows(min_row=2):
            for col in [2, 3]:  # Saturday (C) and Sunday (D)
                games = row[col].value.split(', ') if row[col].value else []
                for game in games:
                    h, a = game.split(' - ')
                    home[h] += 1
                    away[a] += 1
    
        # Print results for all 18 teams
        print(f"{'Abbr':<5} | {'Home':<4} | {'Away':<4} | {'Total':<5} | Status")
        print("-" * 35)
        for team in teams:
            h = home.get(team, 0)
            a = away.get(team, 0)
            status = "✅" if h == 17 and a == 17 else "❌"
            print(f"{team:<5} | {h:<4} | {a:<4} | {h+a:<5} | {status}")
    
    # Run validation
    validate_schedule("Euroleague_Schedule_2025.xlsx")

